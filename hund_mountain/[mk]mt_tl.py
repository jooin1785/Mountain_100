# 기상청 홈페이지 크롤링한 데이터 포함된 엑셀파일 저장


import urllib.request
from bs4 import BeautifulSoup
import numpy as np
import warnings
import openpyxl
import re

warnings.filterwarnings("ignore", category=np.VisibleDeprecationWarning)

filename = 'source/mt_tl.xlsx'
wb = openpyxl.Workbook()
sheet = wb.active

sheet.insert_rows(1)


id_list = ['0009', '0004', '0006', '0012', '0024', '0028', '0040', '0043', '0058', '0059', '0073', '0934', '0091', '0092', '0094', '1254', '0108', '0111', '0112', '0131', '1321', '0138', '0143', '0144', '0149',
           '0152', '0155', '0165', '0167', '0171', '0176', '0217', '0218', '0219', '0225', '0227', '0242', '0247', '0261', '0262', '0263', '0276', '0279', '0281', '1076', '0775', '0317', '0329', '0702', '0370',
           '0383', '0386', '0389', '0397', '0398', '0399', '0422', '0451', '0455', '0456', '0476', '0484', '0492', '0493', '0495', '0505', '0507', '0514', '0522', '0543', '0545', '0548', '0553', '0559', '0573',
           '0575', '0946', '1322', '0591', '0597', '0601', '0606', '0613', '0622', '0623', '0628', '0629', '0638', '0803', '0651', '0652', '0654', '0661', '0679', '0681', '0684', '0686', '0687', '0688', '0699']

number = 1

for id in id_list:

    print(number, id)
    number += 1

    target_url = f'https://www.forest.go.kr/kfsweb/kfi/kfs/foreston/main/contents/ClbngManage/selectMntnInfoDetail.do?mntnId=2000{id}&searchMnt=&searchCnd=10&mn=NKFS_03_01_12&orgId=&mntIndex=1&mntUnit=100'
    resp = urllib.request.urlopen(target_url)
    soup = BeautifulSoup(resp, 'html.parser')

    # 소요시간(M_Time), 난이도(M_level) 저장
    M_Time = soup.select("table.tbl >tbody>tr:nth-child(4)>td:nth-child(2)")[0].text.strip()
    M_level = soup.select("table.tbl >tbody>tr:nth-child(6)>td:nth-child(2)")[0].text.strip()

    sheet.cell(number, 1).value = number-1  #100대명산 id
    sheet.cell(number, 2).value = id  #url요청할때 stnId
    sheet.cell(number, 3).value = M_Time #소요시간
    sheet.cell(number, 4).value = M_level #난이도


    # 선정이유(M_reason), 개관(M_info), 상세설명(M_detail) 저장
    M_Exp = soup.find_all('p')
    p_number = len(M_Exp)
    M_reason, M_view, M_detail = '', '', ''

    match p_number:
        case 2:
            M_reason = str(M_Exp[0].text.strip()) #선정이유
            M_reason = re.sub('<.+?>', '', M_reason, 0).strip()
        case 3:
            M_reason = str(M_Exp[0].text.strip()) #선정이유
            M_reason = re.sub('<.+?>', '', M_reason, 0).strip()
            M_detail = str(M_Exp[1].text.strip()) #상세설명
            M_detail = re.sub('<.+?>', '', M_detail, 0).strip()
        case 4:
            M_reason = str(M_Exp[0].text.strip()) #선정이유
            M_reason = re.sub('<.+?>', '', M_reason, 0).strip()
            M_view = str(M_Exp[1].text.strip()) #개관
            M_view = re.sub('<.+?>', '', M_view, 0).strip()
            M_detail = str(M_Exp[2].text.strip()) #상세설명
            M_detail = re.sub('<.+?>', '', M_detail, 0).strip()
        case 5:
            M_reason = str(M_Exp[1].text.strip()) #선정이유
            M_reason = re.sub('<.+?>', '', M_reason, 0).strip()
            M_detail = str(M_Exp[3].text.strip()) #상세설명
            M_detail = re.sub('<.+?>', '', M_detail, 0).strip()
        case 7:
            M_reason = str(M_Exp[1].text.strip()) #선정이유
            M_reason = re.sub('<.+?>', '', M_reason, 0).strip()
            M_view = str(M_Exp[3].text.strip()) #개관
            M_view = re.sub('<.+?>', '', M_view, 0).strip()
            M_detail = str(M_Exp[5].text.strip()) #상세설명
            M_detail = re.sub('<.+?>', '', M_detail, 0).strip()

    sheet.cell(number, 5).value = M_reason  #선정이유
    sheet.cell(number, 6).value = M_view  #개관
    sheet.cell(number, 7).value = M_detail  #상세설명

wb.save(filename)
