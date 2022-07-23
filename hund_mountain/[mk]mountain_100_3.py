import requests
from xml.etree import ElementTree
import openpyxl
import re
import os
from dotenv import load_dotenv

# api 가져와서 엑셀파일에 저장


filename = 'source/mt_100_3.xlsx'
wb = openpyxl.Workbook()
sheet = wb.active

load_dotenv()
service_key = os.environ.get("DATA_API_KEY")

url = f'http://openapi.forest.go.kr/openapi/service/cultureInfoService/gdTrailInfoOpenAPI?serviceKey={service_key}&numOfRows=100'
resp = requests.get(url)
tree = ElementTree.fromstring(resp.text)
#ElementTree가 resp.text에서 문자열을 가져와 객체로 바꿔줌

#print(tree)
#areatreason, areanm, details, etccourse, mntheight, mntcd, mntnm, overview, subnm, tourisminf, transport
#선정이유, 산정보소재지, 산정보내용, 주변관광정보기타코스, 높이, 산코드, 산명, 산정보개관, 산정보부제, 주변관광정보설명, 대중교통정보
#print(len(tree[1][0])) : items, item의 수
#print(tree[1][0][0][0].text)

id = 0

for item in tree[1][0]:
    id += 1
    sheet.cell(id, 1).value = id
    sheet.cell(id, 2).value = re.sub('<.+?>', '', str(item[0].text), 0).strip() #areatreason -> 선정이유
    sheet.cell(id, 3).value = re.sub('<.+?>', '', str(item[1].text), 0).strip() #areanm
    sheet.cell(id, 4).value = re.sub('<.+?>', '', str(item[2].text), 0).strip() #details -> 상세설명
    sheet.cell(id, 5).value = re.sub('<.+?>', '', str(item[3].text), 0).strip() #etccourse
    sheet.cell(id, 6).value = item[5].text #mntheight
    #sheet.cell(id, 7).value = item[6].text #mntcd
    sheet.cell(id, 7).value = re.sub('<.+?>', '', str(item[7].text), 0).strip() #mntnm -> 산이름
    sheet.cell(id, 8).value = re.sub('<.+?>', '', str(item[8].text), 0).strip() #overview -> 개관
    sheet.cell(id, 9).value = re.sub('<.+?>', '', str(item[9].text), 0).strip() #subnm
    sheet.cell(id, 10).value = re.sub('<.+?>', '', str(item[10].text), 0).strip() #tourisminf
    sheet.cell(id, 11).value = re.sub('<.+?>', '', str(item[11].text), 0).strip() #transport

#가야산, 가리산, 가리왕산 순서 바꿔주기
sheet.insert_rows(1, 1)
sheet.cell(1, 1).value = 1
sheet.cell(1, 2).value = re.sub('<.+?>', '', str(tree[1][0][2][0].text), 0).strip()  # areatreason -> 선정이유
sheet.cell(1, 3).value = re.sub('<.+?>', '', str(tree[1][0][2][1].text), 0).strip()  # areanm
sheet.cell(1, 4).value = re.sub('<.+?>', '', str(tree[1][0][2][2].text), 0).strip()  # details -> 상세설명
sheet.cell(1, 5).value = re.sub('<.+?>', '', str(tree[1][0][2][3].text), 0).strip()  # etccourse
sheet.cell(1, 6).value = tree[1][0][2][5].text  # mntheight
sheet.cell(1, 7).value = re.sub('<.+?>', '', str(tree[1][0][2][7].text), 0).strip()  # mntnm -> 산이름
sheet.cell(1, 8).value = re.sub('<.+?>', '', str(tree[1][0][2][8].text), 0).strip()  # overview -> 개관
sheet.cell(1, 9).value = re.sub('<.+?>', '', str(tree[1][0][2][9].text), 0).strip()  # subnm
sheet.cell(1, 10).value = re.sub('<.+?>', '', str(tree[1][0][2][10].text), 0).strip()  # tourisminf
sheet.cell(1, 11).value = re.sub('<.+?>', '', str(tree[1][0][2][11].text), 0).strip()  # transport
sheet.cell(2, 1).value = 2
sheet.cell(3, 1).value = 3
sheet.delete_rows(4, 1)

#백운산(포천), 백운산(광양), 백운산(정선) 순서 바꿔주기
sheet.insert_rows(43, 1)
sheet.cell(43, 1).value = 43
sheet.cell(43, 2).value = re.sub('<.+?>', '', str(tree[1][0][44][0].text), 0).strip()  # areatreason -> 선정이유
sheet.cell(43, 3).value = re.sub('<.+?>', '', str(tree[1][0][44][1].text), 0).strip()  # areanm
sheet.cell(43, 4).value = re.sub('<.+?>', '', str(tree[1][0][44][2].text), 0).strip()  # details -> 상세설명
sheet.cell(43, 5).value = re.sub('<.+?>', '', str(tree[1][0][44][3].text), 0).strip()  # etccourse
sheet.cell(43, 6).value = tree[1][0][44][5].text  # mntheight
sheet.cell(43, 7).value = re.sub('<.+?>', '', str(tree[1][0][44][7].text), 0).strip()  # mntnm -> 산이름
sheet.cell(43, 8).value = re.sub('<.+?>', '', str(tree[1][0][44][8].text), 0).strip()  # overview -> 개관
sheet.cell(43, 9).value = re.sub('<.+?>', '', str(tree[1][0][44][9].text), 0).strip()  # subnm
sheet.cell(43, 10).value = re.sub('<.+?>', '', str(tree[1][0][44][10].text), 0).strip()  # tourisminf
sheet.cell(43, 11).value = re.sub('<.+?>', '', str(tree[1][0][44][11].text), 0).strip()  # transport
sheet.cell(44, 1).value = 44
sheet.cell(45, 1).value = 45
sheet.delete_rows(46, 1)

sheet.insert_rows(1, 1) #제일 처음에 빈 1행 추가
wb.save(filename)
