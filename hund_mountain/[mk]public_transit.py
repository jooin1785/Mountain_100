from pyproj import Proj, transform
import pandas as pd
import googlemaps
import openpyxl
import os
from dotenv import load_dotenv

load_dotenv()
google_key = os.environ.get("GOOGLE_API_KEY")
gmaps = googlemaps.Client(key=google_key)


# 지하철 위경도 엑셀파일 만들기

subway = pd.read_csv('source/subway.csv', encoding='utf-8') #원본파일 https://www.bigdata-forest.kr/product/PTP002901

filename = 'source/subway_station.xlsx'
wb1 = openpyxl.Workbook()
sheet = wb1.active


WGS84 = {'proj':'latlong', 'datum':'WGS84', 'ellps':'WGS84',} # 구형 좌표계 설정 : EPSG:4326
KATEC = {'proj':'tmerc', 'lat_0':'38N', 'lon_0':'128E', 'ellps':'bessel',
   'x_0':'400000', 'y_0':'600000', 'k':'0.9999', 'a':'6377397.155', 'b':'6356078.9628181886',
   'towgs84':'-115.80,474.99,674.11,1.16,-2.31,-1.63,6.43', 'units':'m'} # 투영 좌표계 설정 : KATEC

inProj = Proj(**KATEC)
outProj = Proj(**WGS84)

for i in range(len(subway)):
   x1, y1 = subway['SBW_STATN_XCRD'][i], subway['SBW_STATN_YCRD'][i]
   x2, y2 = transform(inProj, outProj, x1, y1)
   st_name = subway['SBW_STATN_NM'][i]

   j = i+1
   sheet.cell(j, 1).value = j
   sheet.cell(j, 2).value = st_name
   sheet.cell(j, 3).value = x2
   sheet.cell(j, 4).value = y2

sheet.insert_rows(1, 1)  # 제일 처음에 빈 1행 추가
# no, st_nm, lat, lot

wb1.save(filename)



# 기차역 위경도 엑셀파일 만들기

terminal = pd.read_csv('source/terminal.csv', encoding='utf-8') #원본파일 https://www.bigdata-forest.kr/product/PTP002101

filename = 'source/train_station.xlsx'
wb2 = openpyxl.Workbook()
sheet = wb2.active

j = 0

for i in range(len(terminal)):
    tm_type = int(terminal['BSTP_KIND_TPCD'][i])

    if tm_type == 3:
        j +=1
        sheet.cell(j, 1).value = j
        sheet.cell(j, 2).value = terminal['STATN_NM'][i]
        sheet.cell(j, 3).value = str(terminal['ROAD_NM_ADDR'][i])
        try :
            geocode_result = gmaps.geocode(terminal['ROAD_NM_ADDR'][i])
            sheet.cell(j, 4). value = geocode_result[0]['geometry']['location']['lat']
            sheet.cell(j, 5). value = geocode_result[0]['geometry']['location']['lng']
        except:
            continue
    else:
        continue

    #주소값이 없는 역들 별도로 추가
    addr =''

    for k in [11, 18, 48, 70, 77, 78, 107, 147, 172, 187, 220, 247, 284, 286]:
        match k:
            case 11: #삼산
                addr = '경기도 양평군 양동면 삼산역길 193'
            case 18: #한림정
                addr = '경상남도 김해시 한림면 장방리 1253-188'
            case 48: #진례
                addr = '경상남도 김해시 테크노밸리길 94'
            case 70: #수서
                addr = '서울특별시 강남구 밤고개로 99'
            case 77: #만종
                addr = '강원도 원주시 호저면 운동들2길 21-33'
            case 78: #횡성
                addr = '강원도 횡성군 횡성읍 덕고로 591'
            case 107: #일신
                addr = '경기도 양평군 지평면 노곡길 2'
            case 147: #매곡
                addr = '경기도 양평군 양동면 매월리 1031'
            case 172: #사곡
                addr = '경상북도 구미시 상사서로 173-16'
            case 187: #석불
                addr = '경기도 양평군 지평면 망미리 1319-2'
            case 220: #양원
                addr = '경상북도 봉화군 소천면 분천리 113-2'
            case 247: #운천
                addr = '경기도 파주시 문산읍 운천리 71-7'
            case 284: #진성
                addr = '경상남도 진주시 진성면 진성로 259번길 5-1'
            case 286: #진주수목원
                addr = '경상남도 진주시 일반성면 개암리 144-2'

        geocode_result = gmaps.geocode(addr)
        sheet.cell(k, 3).value = addr
        sheet.cell(k, 4).value = geocode_result[0]['geometry']['location']['lat']
        sheet.cell(k, 5).value = geocode_result[0]['geometry']['location']['lng']

sheet.insert_rows(1, 1)  # 제일 처음에 빈 1행 추가
# no, st_nm, addr, lot, lat

wb2.save(filename)


#버스터미널 위경도 엑셀파일 만들기

bus = pd.read_csv('source/bus.csv', encoding='utf-8')

filename = 'source/bus_station.xlsx'
wb3 = openpyxl.Workbook()
sheet = wb3.active

j = 0

for i in range(len(bus)):
    j += 1
    sheet.cell(j, 1).value = j
    sheet.cell(j, 2).value = bus['tm_nm'][i]
    sheet.cell(j, 3).value = str(bus['addr'][i])
    try :
        geocode_result = gmaps.geocode(bus['addr'][i])
        sheet.cell(j, 4). value = geocode_result[0]['geometry']['location']['lat']
        sheet.cell(j, 5). value = geocode_result[0]['geometry']['location']['lng']
    except:
        continue

sheet.insert_rows(1, 1)  # 제일 처음에 빈 1행 추가
# no, tm_nm, addr, lat, lot

wb3.save(filename)
